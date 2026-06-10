# -*- coding: utf-8 -*-
"""
build_srs_xlsx.py — (re)generate ../srs-system.xlsx from the FMPM curriculum.

WHAT IT DOES
  • Reads data/curriculum_FMPM_S1-S10.txt (the single source of truth).
  • Rebuilds srs-system.xlsx with 7 sheets:
      Dashboard (macro/bird's-eye + charts + exam countdown + progress-over-time)
      Today (daily queue + cheat sheet)   lesson-database (the engine)
      Module View (micro: stats, per-subject table, lesson list)
      Weekly · How-To · History (script-managed daily snapshots)
  • PRESERVES on re-run: study progress (Last Review + Mastery) and Notes
    matched by (Semester | Module | Lesson), the History log, and the exam
    date (Dashboard!J3). Safe to re-run after editing curriculum or visuals.

USAGE   (run from the repo root)
    python3 scripts/build_srs_xlsx.py

⚠️  If you change the column order here, update CONFIG.COL in srs-appscript.js
    too — the Apps Script reads cells by position. And keep the Interval rule
    (column H formula below) identical to the INTERVALS array in the JS.
"""
import re, os
from datetime import date, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(ROOT)
SRC = 'data/curriculum_FMPM_S1-S10.txt'
OUT = 'srs-system.xlsx'
LAST = 2000           # formula range upper bound (data ~1310 rows; headroom to 2000)
H = 2                 # header rows; data starts row 3
IVL = {0: 1, 1: 3, 2: 7, 3: 14, 4: 30, 5: 60}   # mastery -> days (mirror of col H & JS INTERVALS)

# ───────────────────────── parse curriculum ─────────────────────────
def parse(path):
    toks = []
    for raw in open(path, encoding='utf-8'):
        line = raw.rstrip('\n')
        if not line.strip():
            toks.append((-1, '')); continue
        if re.match(r'^S\d+$', line.strip()) and not line.startswith((' ', '-')):
            toks.append((0, line.strip())); continue
        mm = re.match(r'^(\s*)-\s+(.*)$', line)
        if mm:
            toks.append((1 + len(mm.group(1)) // 2, mm.group(2).strip()))
        else:
            toks.append((99, line.strip()))
    out, sem, mod, subj, n = [], None, None, None, len(toks)
    for i, (lvl, txt) in enumerate(toks):
        if lvl == 0:
            sem, mod, subj = txt, None, None
        elif lvl == 1:
            mod, subj = txt, None
        elif lvl == 2:
            j = i + 1
            while j < n and toks[j][0] == -1:
                j += 1
            if j < n and toks[j][0] == 3:
                subj = txt
            else:
                out.append([sem, mod, '', txt]); subj = None
        elif lvl == 3:
            out.append([sem, mod, subj or '', txt])
    return out

rows = parse(SRC)
modpairs = []
for s, m, _, _ in rows:
    if (s, m) not in modpairs:
        modpairs.append((s, m))
NMOD = len(modpairs)
print(f'Parsed {len(rows)} lessons across {NMOD} module/semester pairs')

# ── read existing progress / notes / history / exam date before overwriting ──
progress, hist_rows, exam_date = {}, [], None
if os.path.exists(OUT):
    try:
        old = openpyxl.load_workbook(OUT, data_only=False)
        ex = old['lesson-database']
        for r in range(3, ex.max_row + 1):
            sem, mod, les = ex.cell(r, 2).value, ex.cell(r, 3).value, ex.cell(r, 5).value
            lr, ma = ex.cell(r, 6).value, ex.cell(r, 7).value
            note = ex.cell(r, 14).value if ex.max_column >= 14 else None
            if les and ((lr not in (None, '')) or (ma not in (None, '')) or (note not in (None, ''))):
                progress[(sem, mod, les)] = (lr, ma, note)
        if 'History' in old.sheetnames:
            hs = old['History']
            for r in range(2, hs.max_row + 1):
                vals = [hs.cell(r, c).value for c in range(1, 7)]
                if vals[0] not in (None, ''):
                    hist_rows.append(vals)
        if 'Dashboard' in old.sheetnames:
            j3 = old['Dashboard']['J3'].value
            if j3 not in (None, ''):
                exam_date = j3
        print(f'Preserving: {len(progress)} progressed lessons, {len(hist_rows)} history rows, exam date={exam_date}')
    except Exception as e:
        print('Could not read existing workbook (continuing fresh):', e)

# ───────────────────────── styles ─────────────────────────
WHITE = Font(color='FFFFFF', bold=True)
BOLD = Font(bold=True)
TITLE_FILL = PatternFill('solid', fgColor='1F3864')
HEAD_FILL = PatternFill('solid', fgColor='2E5496')
SUB_FILL = PatternFill('solid', fgColor='D6DCE4')
INPUT_FILL = PatternFill('solid', fgColor='DDEBF7')   # light blue = "edit here"
NOTE_FILL = PatternFill('solid', fgColor='FFF9E3')    # light gold = optional notes
KPI_FILL = PatternFill('solid', fgColor='EAF1FB')
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal='center', vertical='center')
date_fmt = 'yyyy-mm-dd'

def style_header(ws, row, ncols, first_col=1):
    for c in range(first_col, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEAD_FILL; cell.font = WHITE
        cell.alignment = CEN; cell.border = BORDER

def L(col, r1, r2):
    return f"'lesson-database'!${col}${r1}:${col}${r2}"

STATUS_FILLS = {'OVERDUE': 'FFC7CE', 'TODAY': 'C6EFCE', 'TOMORROW': 'BDD7EE',
                'THIS WEEK': 'FFF2CC', 'DONE': 'E2EFDA', 'SCHEDULED': 'F2F2F2'}

def add_status_cf(ws, rng, anchor):
    for kw, color in STATUS_FILLS.items():
        ws.conditional_formatting.add(rng,
            FormulaRule(formula=[f'ISNUMBER(SEARCH("{kw}",{anchor}))'],
                        fill=PatternFill('solid', fgColor=color)))

def add_mastery_cf(ws, rng):
    ws.conditional_formatting.add(rng,
        ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                       mid_type='num', mid_value=2, mid_color='FFEB84',
                       end_type='num', end_value=5, end_color='63BE7B'))

# ───────────────────────── workbook & sheet order ─────────────────────────
wb = Workbook()
db = wb.active; db.title = 'Dashboard'
td = wb.create_sheet('Today')
ldb = wb.create_sheet('lesson-database')
mv = wb.create_sheet('Module View')
wk = wb.create_sheet('Weekly')
ht = wb.create_sheet('How-To')
hist = wb.create_sheet('History')
db.sheet_properties.tabColor = '70AD47'
td.sheet_properties.tabColor = 'C00000'
ldb.sheet_properties.tabColor = '2E75B6'
mv.sheet_properties.tabColor = 'ED7D31'
hist.sheet_properties.tabColor = '999999'

# ═══════════════ lesson-database (engine) ═══════════════
HEADERS = ['#', 'Semester', 'Module', 'Subject', 'Lesson', 'Last Review',
           'Mastery', 'Interval', 'Next Review', 'Status', 'Priority',
           'Synced', 'Event ID', 'Notes']
ldb.merge_cells('A1:N1')
b = ldb['A1']
b.value = ('🧠 MEDICAL SRS — edit only  ▸ F = Last Review (Ctrl+;)  ▸ G = Mastery 0–5  ▸ N = Notes (optional). '
           'Everything else is automatic. Click the ▼ arrows on row 2 to filter by Semester / Module / Subject.')
b.fill = TITLE_FILL; b.font = WHITE; b.alignment = Alignment(horizontal='left', vertical='center')
ldb.row_dimensions[1].height = 26
for c, hdr in enumerate(HEADERS, 1):
    ldb.cell(2, c, hdr)
style_header(ldb, 2, len(HEADERS))
for c in (6, 7):
    ldb.cell(2, c).fill = INPUT_FILL; ldb.cell(2, c).font = Font(bold=True, color='1F3864')
ldb.cell(2, 14).fill = NOTE_FILL; ldb.cell(2, 14).font = Font(bold=True, color='7F6000')

last_data = H + len(rows)
for i, (sem, mod, subj, lesson) in enumerate(rows):
    r = H + 1 + i
    ldb.cell(r, 1, '=ROW()-2').alignment = CEN
    ldb.cell(r, 2, sem); ldb.cell(r, 3, mod); ldb.cell(r, 4, subj); ldb.cell(r, 5, lesson)
    ldb.cell(r, 6).number_format = date_fmt
    ldb.cell(r, 6).fill = INPUT_FILL
    ldb.cell(r, 7).fill = INPUT_FILL; ldb.cell(r, 7).alignment = CEN
    ldb.cell(r, 8, f'=IF($G{r}="",1,IF($G{r}=0,1,IF($G{r}=1,3,IF($G{r}=2,7,IF($G{r}=3,14,IF($G{r}=4,30,60))))))').alignment = CEN
    ldb.cell(r, 9, f'=IF($F{r}="","",$F{r}+$H{r})'); ldb.cell(r, 9).number_format = date_fmt
    ldb.cell(r, 10, f'=IF($F{r}="","⚪ NEW",IF($G{r}>=5,"✅ DONE",IF($I{r}<TODAY(),"🔴 OVERDUE",IF($I{r}=TODAY(),"🟢 TODAY",IF($I{r}=TODAY()+1,"🔵 TOMORROW",IF($I{r}<=TODAY()+7,"📅 THIS WEEK","⏳ SCHEDULED"))))))')
    ldb.cell(r, 11, f'=IF($F{r}="",999,IF($G{r}>=5,9999,IF($I{r}="",999,$I{r}-TODAY())))').alignment = CEN
    ldb.cell(r, 14).fill = NOTE_FILL

# re-apply preserved progress + notes (match on Semester|Module|Lesson)
applied = 0
for i, (sem, mod, subj, lesson) in enumerate(rows):
    hit = progress.get((sem, mod, lesson))
    if hit:
        lr, ma, note = hit; rr = H + 1 + i
        if lr not in (None, ''):
            ldb.cell(rr, 6, lr); ldb.cell(rr, 6).number_format = date_fmt
        if ma not in (None, ''):
            ldb.cell(rr, 7, int(ma))
        if note not in (None, ''):
            ldb.cell(rr, 14, note)
        applied += 1
print(f'Re-applied progress to {applied} lessons')

for col, w in {'A':5,'B':9,'C':26,'D':28,'E':52,'F':13,'G':9,'H':9,'I':13,'J':14,'K':9,'L':9,'M':22,'N':34}.items():
    ldb.column_dimensions[col].width = w
ldb.column_dimensions['M'].hidden = True
ldb.freeze_panes = 'F3'
ldb.auto_filter.ref = f'A2:K{last_data}'
add_mastery_cf(ldb, f'G3:G{last_data}')
add_status_cf(ldb, f'J3:J{last_data}', '$J3')
dv = DataValidation(type='list', formula1='"0,1,2,3,4,5"', allow_blank=True, showErrorMessage=True)
dv.prompt = 'How well did you recall it? 0 = forgot … 5 = mastered'; dv.promptTitle = 'Mastery'
ldb.add_data_validation(dv); dv.add(f'G3:G{last_data}')

# ═══════════════ Today (daily queue + CHEAT SHEET) ═══════════════
td.sheet_view.showGridLines = False
td.merge_cells('A1:J1')
t = td['A1']; t.value = '⚡ TODAY — your study queue & cheat sheet'
t.fill = TITLE_FILL; t.font = Font(color='FFFFFF', bold=True, size=13)
t.alignment = Alignment(horizontal='left', vertical='center'); td.row_dimensions[1].height = 28

quick = [('🔴 Overdue', f'=COUNTIF({L("J",3,LAST)},"*OVERDUE*")'),
         ('🟢 Today',   f'=COUNTIF({L("J",3,LAST)},"*TODAY*")'),
         ('🔵 Tomorrow',f'=COUNTIF({L("J",3,LAST)},"*TOMORROW*")'),
         ('⏱ Est. min', '=(B3+D3)*10')]
for i, (lab, frm) in enumerate(quick):
    c = 1 + i * 2
    td.cell(3, c, lab).fill = KPI_FILL; td.cell(3, c).border = BORDER
    vc = td.cell(3, c + 1, frm); vc.font = Font(bold=True, size=12); vc.alignment = CEN
    vc.fill = KPI_FILL; vc.border = BORDER

td['A5'] = 'MASTERY CHEAT SHEET'; td['A5'].font = BOLD
for c, hh in enumerate(['Score', 'Meaning', 'Comes back in'], 1):
    td.cell(6, c, hh)
style_header(td, 6, 3)
cheat = [(0, 'Forgot completely', '1 day'), (1, 'Very hard', '3 days'),
         (2, 'Shaky', '7 days'), (3, 'OK with effort', '14 days'),
         (4, 'Strong', '30 days'), (5, 'Mastered ✅', 'done — drops off')]
for i, (s, mean, back) in enumerate(cheat):
    r = 7 + i
    td.cell(r, 1, s).alignment = CEN
    td.cell(r, 2, mean); td.cell(r, 3, back)
    for c in range(1, 4):
        td.cell(r, c).border = BORDER
add_mastery_cf(td, 'A7:A12')

td['E5'] = 'GOLDEN RULES'; td['E5'].font = BOLD
rules = ['Type ONLY two cells per review:  F = Last Review (Ctrl+;)  ·  G = Mastery 0–5.',
         'Be honest — "forgot" = 0. The system self-heals; tomorrow it comes back.',
         'Overdue ≠ panic. Work the queue below top-down; the rest reschedules itself.',
         'One module at a time: filter ▼ on lesson-database row 2, or use Module View.']
for i, txt in enumerate(rules):
    td.cell(6 + i, 5, txt)
td['E11'] = 'SHORTCUTS'; td['E11'].font = BOLD
shortcuts = ['Everywhere:  Ctrl+;  = today\'s date   ·   Ctrl+Shift+L = toggle filters (Excel)   ·   Alt+↓ = filter menu',
             'Online palette (Google Sheets):  Ctrl+Alt+Shift+1  → type a lesson → Enter → press 0–5. Done.',
             '   (one-time setup: Extensions → Macros → Import macro → openCommandPalette → assign number 1)']
for i, txt in enumerate(shortcuts):
    td.cell(12 + i, 5, txt)

td['A16'] = 'NOW QUEUE — most overdue first. Study, then log F + G (or use the palette).'
td['A16'].font = BOLD
for c, hh in enumerate(['Sem', 'Module', 'Subject', 'Lesson', 'Last Rev', 'Mast', 'Int', 'Next Rev', 'Status', 'Days'], 1):
    td.cell(17, c, hh)
style_header(td, 17, 10)
td['A18'] = (f"=IFERROR(SORT(FILTER('lesson-database'!$B$3:$K${LAST},'lesson-database'!$K$3:$K${LAST}<=0),10,TRUE),"
             f"\"🎉 Nothing due right now — or this Excel lacks FILTER (use the lesson-database ▼ filters instead).\")")
for r in range(18, 601):
    td.cell(r, 5).number_format = date_fmt
    td.cell(r, 8).number_format = date_fmt
add_mastery_cf(td, 'F18:F600')
add_status_cf(td, 'I18:I600', '$I18')
td.conditional_formatting.add('J18:J600',
    FormulaRule(formula=['AND(ISNUMBER($J18),$J18<0)'], font=Font(color='C00000', bold=True)))
for col, w in {'A':7,'B':24,'C':24,'D':50,'E':12,'F':7,'G':6,'H':12,'I':13,'J':7}.items():
    td.column_dimensions[col].width = w
td.freeze_panes = 'A18'

# ═══════════════ Dashboard (MACRO / bird's-eye) ═══════════════
db.sheet_view.showGridLines = False
db.merge_cells('A1:R1')
t = db['A1']; t.value = "📊 MEDICAL SRS — DASHBOARD  ·  bird's-eye view of every semester & module"
t.fill = TITLE_FILL; t.font = Font(color='FFFFFF', bold=True, size=13)
t.alignment = Alignment(horizontal='left', vertical='center'); db.row_dimensions[1].height = 30

kpis = [
    ('📚 Total lessons',  f"=COUNTA({L('E',3,LAST)})"),
    ('✏️ Started',        f"=COUNTA({L('F',3,LAST)})"),
    ('✅ Mastered',       f"=COUNTIF({L('G',3,LAST)},5)"),
    ('⏳ In progress',    '=B4-B5'),
    ('🔴 Overdue',        f'=COUNTIF({L("J",3,LAST)},"*OVERDUE*")'),
    ('🟢 Due today',      f'=COUNTIF({L("J",3,LAST)},"*TODAY*")'),
    ('📅 Due this week',  f'=COUNTIF({L("J",3,LAST)},"*THIS WEEK*")'),
    ('📈 % Complete',     '=IF(B3=0,0,B5/B3)'),
]
db['A2'] = 'AT A GLANCE'; db['A2'].font = BOLD
for i, (lab, frm) in enumerate(kpis):
    r = 3 + i
    db.cell(r, 1, lab).fill = KPI_FILL
    vc = db.cell(r, 2, frm); vc.font = Font(bold=True, size=12); vc.alignment = CEN; vc.fill = KPI_FILL
    db.cell(r, 1).border = BORDER; vc.border = BORDER
db['B10'].number_format = '0%'

# EXAM COUNTDOWN (J3 = user input, preserved across regenerations)
db.merge_cells('H2:J2'); db['H2'] = '🎯 EXAM COUNTDOWN'; db['H2'].font = BOLD
exam = [('Exam date →', None), ('Days left', '=IF($J$3="","—",$J$3-TODAY())'),
        ('Not yet mastered', '=B3-B5'),
        ('Pace (lessons/day)', '=IF(OR($J$3="",$J$3<=TODAY()),"—",ROUNDUP((B3-B5)/($J$3-TODAY()),1))')]
for i, (lab, frm) in enumerate(exam):
    r = 3 + i
    db.merge_cells(f'H{r}:I{r}')
    db.cell(r, 8, lab).border = BORDER
    vc = db.cell(r, 10)
    if frm: vc.value = frm
    vc.font = Font(bold=True, size=12); vc.alignment = CEN; vc.border = BORDER
db['J3'].fill = INPUT_FILL; db['J3'].number_format = date_fmt
if exam_date:
    db['J3'] = exam_date; db['J3'].number_format = date_fmt

db['A12'] = 'BY SEMESTER'; db['A12'].font = BOLD
for c, hh in enumerate(['Semester', 'Total', 'Started', 'Mastered', 'Overdue', '% Done'], 1):
    db.cell(13, c, hh)
style_header(db, 13, 6)
for i in range(10):
    r = 14 + i; s = f'S{i+1}'
    db.cell(r, 1, s).alignment = CEN
    db.cell(r, 2, f"=COUNTIF({L('B',3,LAST)},$A{r})")
    db.cell(r, 3, f'=COUNTIFS({L("B",3,LAST)},$A{r},{L("F",3,LAST)},"<>")')
    db.cell(r, 4, f"=COUNTIFS({L('B',3,LAST)},$A{r},{L('G',3,LAST)},5)")
    db.cell(r, 5, f'=COUNTIFS({L("B",3,LAST)},$A{r},{L("J",3,LAST)},"*OVERDUE*")')
    db.cell(r, 6, f'=IF($B{r}=0,0,$D{r}/$B{r})'); db.cell(r, 6).number_format = '0%'
    for c in range(1, 7):
        db.cell(r, c).border = BORDER
sem_last = 23
db.conditional_formatting.add(f'F14:F{sem_last}',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color='63BE7B'))

db['H12'] = 'MASTERY MIX'; db['H12'].font = BOLD
db.cell(13, 8, 'Mastery'); db.cell(13, 9, 'Lessons')
style_header(db, 13, 9, first_col=8)
for k in range(6):
    r = 14 + k
    db.cell(r, 8, k).alignment = CEN
    db.cell(r, 9, f"=COUNTIF({L('G',3,LAST)},$H{r})")
    db.cell(r, 8).border = BORDER; db.cell(r, 9).border = BORDER

for col, w in {'A':20,'B':9,'C':9,'D':10,'E':9,'F':9,'G':3,'H':9,'I':9,'J':11}.items():
    db.column_dimensions[col].width = w

c1 = BarChart(); c1.type = 'col'; c1.title = 'Lessons per semester'; c1.legend = None
c1.add_data(Reference(db, min_col=2, min_row=13, max_row=sem_last), titles_from_data=True)
c1.set_categories(Reference(db, min_col=1, min_row=14, max_row=sem_last))
c1.height, c1.width = 7.5, 15; db.add_chart(c1, 'K3')
c2 = BarChart(); c2.type = 'col'; c2.title = 'Progress per semester'; c2.grouping = 'stacked'; c2.overlap = 100
c2.add_data(Reference(db, min_col=3, max_col=4, min_row=13, max_row=sem_last), titles_from_data=True)
c2.set_categories(Reference(db, min_col=1, min_row=14, max_row=sem_last))
c2.height, c2.width = 7.5, 15; db.add_chart(c2, 'K19')
c3 = BarChart(); c3.type = 'col'; c3.title = 'Mastery distribution (0=forgot → 5=mastered)'; c3.legend = None
c3.add_data(Reference(db, min_col=9, min_row=13, max_row=19), titles_from_data=True)
c3.set_categories(Reference(db, min_col=8, min_row=14, max_row=19))
c3.height, c3.width = 7.5, 15; db.add_chart(c3, 'A26')
cH = LineChart(); cH.title = 'Progress over time (auto-logged daily when online)'
cH.add_data(Reference(hist, min_col=3, max_col=5, min_row=1, max_row=400), titles_from_data=True)
cH.set_categories(Reference(hist, min_col=1, min_row=2, max_row=400))
cH.height, cH.width = 7.5, 15; db.add_chart(cH, 'K35')

MHR = 45
db.cell(MHR - 1, 1, 'BY MODULE  (pick any of these in Module View, or filter lesson-database)').font = BOLD
for c, hh in enumerate(['Module (pick this in Module View)', 'Sem', 'Module name', 'Total', 'Started', 'Mastered', 'Overdue', '% Done'], 1):
    db.cell(MHR, c, hh)
style_header(db, MHR, 8)
for i, (s, m) in enumerate(modpairs):
    r = MHR + 1 + i
    db.cell(r, 1, f'{s} — {m}')
    db.cell(r, 2, s).alignment = CEN
    db.cell(r, 3, m)
    db.cell(r, 4, f"=COUNTIFS({L('B',3,LAST)},$B{r},{L('C',3,LAST)},$C{r})")
    db.cell(r, 5, f'=COUNTIFS({L("B",3,LAST)},$B{r},{L("C",3,LAST)},$C{r},{L("F",3,LAST)},"<>")')
    db.cell(r, 6, f"=COUNTIFS({L('B',3,LAST)},$B{r},{L('C',3,LAST)},$C{r},{L('G',3,LAST)},5)")
    db.cell(r, 7, f'=COUNTIFS({L("B",3,LAST)},$B{r},{L("C",3,LAST)},$C{r},{L("J",3,LAST)},"*OVERDUE*")')
    db.cell(r, 8, f'=IF($D{r}=0,0,$F{r}/$D{r})'); db.cell(r, 8).number_format = '0%'
    for c in range(1, 9):
        db.cell(r, c).border = BORDER
mod_last = MHR + NMOD
db.conditional_formatting.add(f'H{MHR+1}:H{mod_last}',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color='5B9BD5'))
c4 = BarChart(); c4.type = 'bar'; c4.title = 'Lessons per module'; c4.legend = None
c4.add_data(Reference(db, min_col=4, min_row=MHR, max_row=mod_last), titles_from_data=True)
c4.set_categories(Reference(db, min_col=1, min_row=MHR + 1, max_row=mod_last))
c4.height, c4.width = 28, 16; db.add_chart(c4, 'K52')

# ═══════════════ Module View (MICRO) ═══════════════
mv.sheet_view.showGridLines = False
mv.merge_cells('A1:H1')
t = mv['A1']; t.value = '🔍 MODULE VIEW — pick one module: stats, subjects, lessons, mastery chart'
t.fill = TITLE_FILL; t.font = Font(color='FFFFFF', bold=True, size=12)
t.alignment = Alignment(horizontal='left', vertical='center'); mv.row_dimensions[1].height = 26
mv['A3'] = '▶ Module:'; mv['A3'].font = Font(bold=True, size=12)
mv.merge_cells('C3:F3')
pick = mv['C3']; pick.value = f'{modpairs[0][0]} — {modpairs[0][1]}'
pick.fill = PatternFill('solid', fgColor='FCE4D6'); pick.font = Font(bold=True, size=12)
pick.alignment = CEN; pick.border = BORDER
dvm = DataValidation(type='list', formula1=f'=Dashboard!$A${MHR+1}:$A${mod_last}', allow_blank=False, showErrorMessage=True)
dvm.prompt = 'Choose the module to inspect'; dvm.promptTitle = 'Module'
mv.add_data_validation(dvm); dvm.add('C3')
mv['N1'] = f'=IFERROR(INDEX(Dashboard!$B${MHR+1}:$B${mod_last},MATCH($C$3,Dashboard!$A${MHR+1}:$A${mod_last},0)),"")'
mv['N2'] = f'=IFERROR(INDEX(Dashboard!$C${MHR+1}:$C${mod_last},MATCH($C$3,Dashboard!$A${MHR+1}:$A${mod_last},0)),"")'
mv.column_dimensions['N'].hidden = True

def cifs(extra=''):
    return f"=COUNTIFS({L('B',3,LAST)},$N$1,{L('C',3,LAST)},$N$2{extra})"

stats = [('Total', cifs()), ('Started', cifs(f',{L("F",3,LAST)},"<>"')),
         ('Mastered', cifs(f',{L("G",3,LAST)},5')), ('Overdue', cifs(f',{L("J",3,LAST)},"*OVERDUE*"')),
         ('Due today', cifs(f',{L("J",3,LAST)},"*TODAY*"')), ('This week', cifs(f',{L("J",3,LAST)},"*THIS WEEK*"'))]
mv['A5'] = 'THIS MODULE'; mv['A5'].font = BOLD
for i, (lab, frm) in enumerate(stats):
    c = 1 + (i % 3) * 2; r = 6 + i // 3
    mv.cell(r, c, lab).fill = SUB_FILL; mv.cell(r, c).border = BORDER
    vc = mv.cell(r, c + 1, frm); vc.font = BOLD; vc.alignment = CEN; vc.border = BORDER

mv['H5'] = 'Mastery'; mv['I5'] = 'Lessons'
style_header(mv, 5, 9, first_col=8)
for k in range(6):
    r = 6 + k
    mv.cell(r, 8, k).alignment = CEN; mv.cell(r, 8).border = BORDER
    mv.cell(r, 9, cifs(f',{L("G",3,LAST)},$H{r}')); mv.cell(r, 9).border = BORDER
cm = BarChart(); cm.type = 'col'; cm.title = 'Mastery in this module'; cm.legend = None
cm.add_data(Reference(mv, min_col=9, min_row=5, max_row=11), titles_from_data=True)
cm.set_categories(Reference(mv, min_col=8, min_row=6, max_row=11))
cm.height, cm.width = 6.5, 12; mv.add_chart(cm, 'K5')

mv['A9'] = 'BY SUBJECT (needs Excel 365 / Google Sheets)'; mv['A9'].font = BOLD
for c, hh in enumerate(['Subject', 'Total', 'Started', 'Mastered', '% Done'], 1):
    mv.cell(10, c, hh)
style_header(mv, 10, 5)
for i in range(12):
    r = 11 + i
    mv.cell(r, 1, f"=IFERROR(INDEX(UNIQUE(FILTER({L('D',3,LAST)},({L('B',3,LAST)}=$N$1)*({L('C',3,LAST)}=$N$2)*({L('D',3,LAST)}<>\"\"))),ROW()-10),\"\")")
    mv.cell(r, 2, f'=IF($A{r}="","",COUNTIFS({L("B",3,LAST)},$N$1,{L("C",3,LAST)},$N$2,{L("D",3,LAST)},$A{r}))')
    mv.cell(r, 3, f'=IF($A{r}="","",COUNTIFS({L("B",3,LAST)},$N$1,{L("C",3,LAST)},$N$2,{L("D",3,LAST)},$A{r},{L("F",3,LAST)},"<>"))')
    mv.cell(r, 4, f'=IF($A{r}="","",COUNTIFS({L("B",3,LAST)},$N$1,{L("C",3,LAST)},$N$2,{L("D",3,LAST)},$A{r},{L("G",3,LAST)},5))')
    mv.cell(r, 5, f'=IF(OR($A{r}="",$B{r}=0),"",$D{r}/$B{r})'); mv.cell(r, 5).number_format = '0%'
mv.conditional_formatting.add('E11:E22',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color='ED7D31'))

mv['A24'] = 'LESSONS IN THIS MODULE'; mv['A24'].font = BOLD
for c, hh in enumerate(['Lesson', 'Last Review', 'Mastery', 'Interval', 'Next Review', 'Status', 'Priority'], 1):
    mv.cell(25, c, hh)
style_header(mv, 25, 7)
block = f"'lesson-database'!$E$3:$K${LAST}"
mv['A26'] = (f"=IFERROR(FILTER({block},({L('B',3,LAST)}=$N$1)*({L('C',3,LAST)}=$N$2),\"— no lessons —\"),"
             f"\"FILTER needs Excel 365 or Google Sheets — use the lesson-database tab's filter arrows instead.\")")
for r in range(26, 601):
    mv.cell(r, 2).number_format = date_fmt
    mv.cell(r, 5).number_format = date_fmt
add_mastery_cf(mv, 'C26:C600')
add_status_cf(mv, 'F26:F600', '$F26')
for col, w in {'A':50,'B':13,'C':9,'D':10,'E':13,'F':14,'G':9,'H':9,'I':9}.items():
    mv.column_dimensions[col].width = w
mv.freeze_panes = 'A26'

# ═══════════════ Weekly ═══════════════
wk.sheet_view.showGridLines = False
wk.merge_cells('A1:E1')
t = wk['A1']; t.value = '🗓️ WEEKLY — your next 7 days of reviews'
t.fill = TITLE_FILL; t.font = WHITE; t.alignment = Alignment(horizontal='left', vertical='center')
wk.row_dimensions[1].height = 24
for c, hh in enumerate(['Date', 'Day', 'Reviews', 'Load', 'Est. min'], 1):
    wk.cell(3, c, hh)
style_header(wk, 3, 5)
for i in range(7):
    r = 4 + i
    wk.cell(r, 1, '=TODAY()' if i == 0 else f'=A{r-1}+1'); wk.cell(r, 1).number_format = date_fmt
    wk.cell(r, 2, f'=TEXT(A{r},"ddd")').alignment = CEN
    if i == 0:
        wk.cell(r, 3, f'=COUNTIF({L("I",3,LAST)},A{r})+COUNTIF({L("J",3,LAST)},"*OVERDUE*")')
    else:
        wk.cell(r, 3, f'=COUNTIF({L("I",3,LAST)},A{r})')
    wk.cell(r, 3).alignment = CEN
    wk.cell(r, 4, f'=IF(C{r}=0,"🟢 Free",IF(C{r}<=5,"🟢 Light",IF(C{r}<=10,"🟡 Medium",IF(C{r}<=15,"🟠 Heavy","🔴 Overload"))))')
    wk.cell(r, 5, f'=C{r}*10').alignment = CEN
    for c in range(1, 6):
        wk.cell(r, c).border = BORDER
for col, w in {'A':13,'B':7,'C':10,'D':13,'E':10}.items():
    wk.column_dimensions[col].width = w
cw = BarChart(); cw.type = 'col'; cw.title = 'Reviews per day'; cw.legend = None
cw.add_data(Reference(wk, min_col=3, min_row=3, max_row=10), titles_from_data=True)
cw.set_categories(Reference(wk, min_col=2, min_row=4, max_row=10))
cw.height, cw.width = 7, 13; wk.add_chart(cw, 'G3')

# ═══════════════ How-To ═══════════════
ht.sheet_view.showGridLines = False
ht.column_dimensions['A'].width = 112
guide = [
    ('🧠 MEDICAL SRS — HOW TO USE', True),
    ('', False),
    ('THE ONE RULE: per review you type two cells on lesson-database —', True),
    ('   • F  Last Review  → the day you studied it (press Ctrl+;  for today)', False),
    ('   • G  Mastery      → how well you recalled it, 0 (forgot) to 5 (mastered)', False),
    ('   (• N Notes is optional — mnemonics, page refs, whatever helps.)', False),
    ('Everything else (interval, next review, status, priority) fills in by itself.', False),
    ('', False),
    ('THE SCHEDULE: mastery 0→1d · 1→3d · 2→7d · 3→14d · 4→30d · 5→done (drops off).', True),
    ('', False),
    ('THE TABS', True),
    ('   • Today       — START HERE every day: queue of what is due + the cheat sheet.', False),
    ("   • Dashboard   — bird's-eye: KPIs, exam countdown, per-semester/module tables, 5 charts.", False),
    ('   • lesson-database — the engine; the only place you type.', False),
    ('   • Module View — pick one module: stats, per-subject breakdown, lessons, chart.', False),
    ('   • Weekly      — workload for the next 7 days.', False),
    ('   • History     — script-managed daily snapshots (do NOT edit; feeds the progress chart).', False),
    ('', False),
    ('FILTERING (one module / one semester)', True),
    ('   Fast & works everywhere: lesson-database → ▼ arrow on row 2 → pick Semester or Module.', False),
    ('   Deep view: Module View dropdown.   Online: open the palette and just type the module name.', False),
    ('', False),
    ('SHORTCUTS', True),
    ("   Everywhere : Ctrl+;  inserts today's date · Ctrl+Shift+L toggles filters (Excel) · Alt+↓ filter menu.", False),
    ('   Online command palette (Google Sheets only — like Ctrl+K):', False),
    ('     One-time: Extensions → Macros → Import macro → openCommandPalette;', False),
    ('               then Extensions → Macros → Manage macros → assign it number 1.', False),
    ('     Use: Ctrl+Alt+Shift+1 → type lesson/module/command → Enter → press 0–5 to log mastery.', False),
    ('     (Sheets reserves the real Ctrl+K for itself; inside the palette Ctrl+K refocuses search.)', False),
    ('', False),
    ('ONLINE vs OFFLINE', True),
    ('   • Online  (Google Sheets): + calendar sync, command palette, auto snapshots (📚 SRS menu).', False),
    ('   • Offline (Microsoft Excel): everything formula-based works; FILTER lists need Excel 365.', False),
    ('', False),
    ('MAINTENANCE', True),
    ('   Add/rename lessons in data/curriculum_FMPM_S1-S10.txt, then run:', False),
    ('       python3 scripts/build_srs_xlsx.py', False),
    ('   It regenerates this workbook and keeps your progress, notes, history and exam date.', False),
]
for i, (txt, bold) in enumerate(guide):
    cell = ht.cell(i + 1, 1, txt)
    cell.font = Font(bold=bold, size=13 if i == 0 else 11)
    if i == 0:
        cell.fill = TITLE_FILL; cell.font = Font(color='FFFFFF', bold=True, size=13)

# ═══════════════ History (script-managed) ═══════════════
for c, hh in enumerate(['Date', 'Total', 'Started', 'Mastered', 'Overdue', 'Reviewed that day'], 1):
    hist.cell(1, c, hh)
style_header(hist, 1, 6)
hist['H1'] = '⚠️ Script-managed — do not edit. Apps Script appends one row per day (online); feeds the Dashboard progress chart.'
hist['H1'].font = Font(italic=True, color='808080')
if hist_rows:
    for i, vals in enumerate(hist_rows):
        for c, v in enumerate(vals, 1):
            hist.cell(2 + i, c, v)
        hist.cell(2 + i, 1).number_format = date_fmt
else:
    today = date.today()
    started = sum(1 for v in progress.values() if v[0] not in (None, ''))
    mastered = sum(1 for v in progress.values() if v[1] not in (None, '') and int(v[1]) >= 5)
    overdue = 0
    for lr, ma, _ in progress.values():
        if lr in (None, ''):
            continue
        m = int(ma) if ma not in (None, '') else 0
        if m >= 5:
            continue
        d = lr.date() if hasattr(lr, 'date') else lr
        if d + timedelta(days=IVL.get(m, 1)) < today:
            overdue += 1
    hist.append([today, len(rows), started, mastered, overdue, 0])
    hist.cell(2, 1).number_format = date_fmt
for col, w in {'A':12,'B':8,'C':8,'D':10,'E':9,'F':16}.items():
    hist.column_dimensions[col].width = w

try:
    wb.calculation.fullCalcOnLoad = True
except Exception as e:
    print('calc flag skip:', e)
wb.active = 1   # Today
wb.save(OUT)
print(f'Saved {OUT} · {os.path.getsize(OUT)} bytes · sheets {wb.sheetnames} · data rows 3..{last_data}')
